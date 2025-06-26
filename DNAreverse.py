import openpyxl

# 计算DNA序列的反向互补
def reverse_complement(dna_sequence):
    complement = {'A': 'T', 'T': 'A', 'C': 'G', 'G': 'C'}
    return ''.join(complement[base] for base in reversed(dna_sequence))

# 打开xlsx文件
wb = openpyxl.load_workbook('/Users/stepviewmaifu/Downloads/test.xlsx')  # 请替换成你的文件名
ws = wb.active

# 假设DNA序列在第一列，反向互补输出在第二列
for row in range(1, ws.max_row + 1):
    dna_sequence = ws.cell(row=row, column=1).value  # 获取第一列的DNA序列
    if dna_sequence:  # 确保该单元格不为空
        reversed_complement = reverse_complement(dna_sequence)
        ws.cell(row=row, column=2).value = reversed_complement  # 将反向互补结果写入第二列

# 保存修改后的文件
wb.save('/Users/stepviewmaifu/Downloads/testxlsx.xlsx')  # 生成新文件